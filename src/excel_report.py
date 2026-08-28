"""Formula-driven Excel workbook for the Windansea Coconuts State of the Union.

This is the audit artifact and accountant hand-off. CRITICAL: every total and
average is written as a live Excel FORMULA (SUMIFS, SUMPRODUCT, AVERAGE), never
a baked Python number, so the workbook recalculates when someone edits the
adjustable window or fixes a transaction. Bucket and flag strings come from
src.constants so the formula criteria always match the data.

No dash characters are used as pauses in any label or note. Commas, periods,
and parentheses only. (Formulas may use "-" for arithmetic negation.)
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from src import store
from src import constants as k

# Transactions columns: A..J. Amount=F, Bucket=G, Flag=H, Month=J.
TXN_HEADERS = ["Date", "Source", "Account", "Counterparty", "Category",
               "Amount", "Bucket", "Flag", "Review", "Month"]
MONTHS = ["2026-01", "2026-02", "2026-03", "2026-04", "2026-05", "2026-06"]
CURRENCY_FMT = "#,##0.00"
HEADER_FONT = Font(name="Calibri", size=11, bold=True)
BASE_FONT = Font(name="Calibri", size=11)
TITLE_FONT = Font(name="Calibri", size=11, bold=True)


def _txn_col(letter):
    return f"Transactions!${letter}:${letter}"


def _build_transactions(ws, conn):
    """Write the Transactions tab and define it as an Excel Table 'Txns'."""
    for col, head in enumerate(TXN_HEADERS, start=1):
        c = ws.cell(row=1, column=col, value=head)
        c.font = HEADER_FONT

    rows = conn.execute(
        "SELECT date, source, account, counterparty, category, amount, "
        "bucket, flag, review FROM transactions ORDER BY date ASC"
    ).fetchall()

    for i, r in enumerate(rows, start=2):
        date = r[0] or ""
        ws.cell(row=i, column=1, value=date).font = BASE_FONT
        ws.cell(row=i, column=2, value=r[1]).font = BASE_FONT
        ws.cell(row=i, column=3, value=r[2]).font = BASE_FONT
        ws.cell(row=i, column=4, value=r[3]).font = BASE_FONT
        ws.cell(row=i, column=5, value=r[4]).font = BASE_FONT
        amt = ws.cell(row=i, column=6, value=r[5])
        amt.font = BASE_FONT
        amt.number_format = CURRENCY_FMT
        ws.cell(row=i, column=7, value=r[6]).font = BASE_FONT
        ws.cell(row=i, column=8, value=r[7]).font = BASE_FONT
        ws.cell(row=i, column=9, value=r[8]).font = BASE_FONT
        # Month key = first 7 chars of the date, literal text.
        ws.cell(row=i, column=10, value=date[:7]).font = BASE_FONT

    last_row = max(2, len(rows) + 1)
    ref = f"A1:{get_column_letter(len(TXN_HEADERS))}{last_row}"
    table = Table(displayName="Txns", ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(table)

    widths = [12, 12, 12, 24, 18, 14, 18, 30, 10, 10]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w


def _sumifs_bucket(bucket, month_cell=None, negate=False):
    """Build a SUMIFS formula on Amount filtered by Bucket (and Month)."""
    base = f'SUMIFS({_txn_col("F")}, {_txn_col("G")}, "{bucket}"'
    if month_cell is not None:
        base += f", {_txn_col('J')}, {month_cell}"
    base += ")"
    return f"=-{base}" if negate else f"={base}"


def _set(ws, row, col, value, font=None, fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font or BASE_FONT
    if fmt:
        c.number_format = fmt
    return c


def _build_monthly_summary(ws, conn):
    """Monthly grid, adjustable window averages, State of the Union, Owner Comp."""
    ws["A1"] = "Monthly Summary"
    ws["A1"].font = TITLE_FONT

    # Row 3: month label row that the SUMIFS formulas reference (text).
    label_row = 3
    _set(ws, label_row, 1, "Month", HEADER_FONT)
    first_data_col = 2
    for j, month in enumerate(MONTHS):
        col = first_data_col + j
        _set(ws, label_row, col, month, HEADER_FONT)

    last_data_col = first_data_col + len(MONTHS) - 1
    last_col_letter = get_column_letter(last_data_col)
    first_col_letter = get_column_letter(first_data_col)

    def month_cell(j):
        # Absolute reference to the month label in the label row.
        return f"${get_column_letter(first_data_col + j)}${label_row}"

    # Metric rows. Each entry: (label, bucket, negate).
    metric_rows = [
        ("Revenue", k.BUCKET_REVENUE, False),
        ("Expenses", k.BUCKET_OPERATING, True),
        ("Payroll", k.BUCKET_PAYROLL, True),
        ("Owner Pay", k.BUCKET_OWNER_PAY, True),
        ("Owner Draw", k.BUCKET_OWNER_DRAW, True),
    ]
    row_of = {}
    r = label_row + 1
    for label, bucket, negate in metric_rows:
        _set(ws, r, 1, label, HEADER_FONT)
        for j in range(len(MONTHS)):
            col = first_data_col + j
            f = _sumifs_bucket(bucket, month_cell=month_cell(j), negate=negate)
            _set(ws, r, col, f, fmt=CURRENCY_FMT)
        row_of[label] = r
        r += 1

    # Net row = Revenue - Expenses - Payroll - Owner Pay - Owner Draw (per col).
    net_row = r
    _set(ws, net_row, 1, "Net", HEADER_FONT)
    for j in range(len(MONTHS)):
        col = first_data_col + j
        cl = get_column_letter(col)
        f = (f"={cl}{row_of['Revenue']}-{cl}{row_of['Expenses']}"
             f"-{cl}{row_of['Payroll']}-{cl}{row_of['Owner Pay']}"
             f"-{cl}{row_of['Owner Draw']}")
        _set(ws, net_row, col, f, fmt=CURRENCY_FMT)
    row_of["Net"] = net_row

    # Adjustable window input cells.
    win_row = net_row + 2
    _set(ws, win_row, 1, "Window Start", HEADER_FONT)
    start_cell = f"$B${win_row}"
    _set(ws, win_row, 2, "2026-04")
    _set(ws, win_row + 1, 1, "Window End", HEADER_FONT)
    end_cell = f"$B${win_row + 1}"
    _set(ws, win_row + 1, 2, "2026-06")

    # Averages via SUMPRODUCT so they recalc with the window cells.
    mlab = f"${first_col_letter}${label_row}:${last_col_letter}${label_row}"

    def avg_formula(values_row):
        vrow = f"${first_col_letter}${values_row}:${last_col_letter}${values_row}"
        num = (f"SUMPRODUCT(({mlab}>={start_cell})*({mlab}<={end_cell})*{vrow})")
        den = (f"SUMPRODUCT(({mlab}>={start_cell})*({mlab}<={end_cell})*1)")
        return f"={num}/{den}"

    avg_row = win_row + 3
    _set(ws, avg_row, 1, "Revenue Avg (window)", HEADER_FONT)
    _set(ws, avg_row, 2, avg_formula(row_of["Revenue"]), fmt=CURRENCY_FMT)
    _set(ws, avg_row + 1, 1, "Expenses Avg (window)", HEADER_FONT)
    _set(ws, avg_row + 1, 2, avg_formula(row_of["Expenses"]), fmt=CURRENCY_FMT)
    _set(ws, avg_row + 2, 1, "Net Avg (window)", HEADER_FONT)
    _set(ws, avg_row + 2, 2, avg_formula(row_of["Net"]), fmt=CURRENCY_FMT)

    # State of the Union block (formula-driven YTD).
    sotu_row = avg_row + 4
    _set(ws, sotu_row, 1, "State of the Union (Year to Date)", TITLE_FONT)
    sotu = [
        ("YTD Revenue", _sumifs_bucket(k.BUCKET_REVENUE)),
        ("YTD Expenses", _sumifs_bucket(k.BUCKET_OPERATING, negate=True)),
        ("Total Payroll", _sumifs_bucket(k.BUCKET_PAYROLL, negate=True)),
        ("Total Owner Pay", _sumifs_bucket(k.BUCKET_OWNER_PAY, negate=True)),
        ("Total Owner Draws", _sumifs_bucket(k.BUCKET_OWNER_DRAW, negate=True)),
        ("YTD Net",
         f'=SUMIFS({_txn_col("F")}, {_txn_col("G")}, "<>{k.BUCKET_EXCLUDED}")'),
    ]
    rr = sotu_row + 1
    for label, formula in sotu:
        _set(ws, rr, 1, label, HEADER_FONT)
        _set(ws, rr, 2, formula, fmt=CURRENCY_FMT)
        rr += 1

    # Owner Comp block.
    oc_row = rr + 1
    _set(ws, oc_row, 1, "Owner Compensation", TITLE_FONT)
    rr = oc_row + 1

    car_row = rr
    _set(ws, car_row, 1, "Jordan Capital One car", HEADER_FONT)
    _set(ws, car_row, 2,
         f'=-SUMIFS({_txn_col("F")}, {_txn_col("H")}, "{k.FLAG_OWNER_CAR}")',
         fmt=CURRENCY_FMT)
    rr += 1

    bofa_row = rr
    _set(ws, bofa_row, 1, "Jordan BofA 2k draw", HEADER_FONT)
    _set(ws, bofa_row, 2,
         f'=-SUMIFS({_txn_col("F")}, {_txn_col("H")}, "{k.FLAG_OWNER_DRAW_BOFA}")',
         fmt=CURRENCY_FMT)
    rr += 1

    jordan_total_row = rr
    _set(ws, jordan_total_row, 1, "Jordan total", HEADER_FONT)
    _set(ws, jordan_total_row, 2, f"=B{car_row}+B{bofa_row}", fmt=CURRENCY_FMT)
    rr += 1

    direct_row = rr
    _set(ws, direct_row, 1, "Harrison direct transfers", HEADER_FONT)
    _set(ws, direct_row, 2,
         f'=-SUMIFS({_txn_col("F")}, {_txn_col("H")}, "{k.FLAG_HARRISON_DIRECT}")',
         fmt=CURRENCY_FMT)
    rr += 1

    # Harrison ADP W-2 is a VALUE from meta (not in the transactions table).
    raw_w2 = store.get_meta(conn, k.META_HARRISON_W2_GROSS)
    try:
        adp_w2 = float(raw_w2) if raw_w2 is not None else 0.0
    except (TypeError, ValueError):
        adp_w2 = 0.0
    adp_row = rr
    _set(ws, adp_row, 1, "Harrison ADP W-2", HEADER_FONT)
    _set(ws, adp_row, 2, adp_w2, fmt=CURRENCY_FMT)
    _set(ws, adp_row, 3, "(from ADP file)")
    rr += 1

    harrison_total_row = rr
    _set(ws, harrison_total_row, 1, "Harrison total", HEADER_FONT)
    _set(ws, harrison_total_row, 2, f"=B{adp_row}+B{direct_row}", fmt=CURRENCY_FMT)
    rr += 1

    # Notes (no dashes used as pauses).
    rr += 1
    _set(ws, rr, 1,
         "Note. ADP W-2 wages are also included in the Payroll total above. "
         "Direct transfers are separate.")
    rr += 1
    _set(ws, rr, 1,
         "Note. The two Jordan car items are distinct. The Capital One amount "
         "is the car loan paid via the business. The BofA item is a 2k owner "
         "draw toward a separate car.")

    ws.column_dimensions["A"].width = 34
    for j in range(len(MONTHS) + 1):
        ws.column_dimensions[get_column_letter(first_data_col + j)].width = 14


def build(conn, out_path):
    """Build the workbook at out_path. Returns the path."""
    wb = Workbook()
    txn_ws = wb.active
    txn_ws.title = "Transactions"
    _build_transactions(txn_ws, conn)

    ms_ws = wb.create_sheet("Monthly Summary")
    _build_monthly_summary(ms_ws, conn)

    wb.save(out_path)
    return out_path
