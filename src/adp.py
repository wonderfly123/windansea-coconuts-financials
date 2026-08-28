"""ADP payroll summary parser for the Windansea Coconuts financial dashboard.

Parses the ADP "Payroll Summary" workbook export into normalized per-pay-run
rows plus company totals, and reconciles those totals against the figures the
report itself prints so downstream code can trust the parse.

Documented ground-truth checksums (Jan 23 - Jun 15, 2026 window):
    gross paid          : 39,330.80
    employee tax        :  5,760.16  (sits INSIDE gross, not additional cash)
    employer liability  :  4,094.80
    total cash cost     : 43,425.60
    cash leaving bank   : gross + employer liability = 43,425.60
"""

from typing import Any, Dict, List, Optional

import openpyxl

# Header labels as they appear in the ADP export, mapped to our field names.
_COLUMN_LABELS = {
    "Check Date": "date",
    "Name": "employee",
    "Hours": "hours",
    "Total Paid": "gross",
    "Tax Withheld": "employee_tax",
    "Deductions": "deductions",
    "Net Pay": "net_pay",
    "Employer Liability": "employer_liability",
    "Total Expenses": "total_expense",
}

# Numeric fields we coerce to float.
_NUMERIC_FIELDS = (
    "hours",
    "gross",
    "employee_tax",
    "deductions",
    "net_pay",
    "employer_liability",
    "total_expense",
)


def _to_float(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    return float(value)


def _find_header(ws) -> Optional[Dict[str, int]]:
    """Locate the header row and return {field_name: column_index}."""
    for row in ws.iter_rows():
        labels = {str(c.value).strip(): c.column - 1 for c in row if c.value is not None}
        if "Total Paid" in labels and "Name" in labels:
            return {
                field: labels[label]
                for label, field in _COLUMN_LABELS.items()
                if label in labels
            }
    return None


def load_adp(path: str) -> Dict[str, Any]:
    """Parse the ADP payroll summary workbook.

    Returns a dict with:
      - "rows": list of per-pay-run/per-employee records, each with at least
        date, employee, gross, employee_tax, employer_liability.
      - "totals": gross, employee_tax, employer_liability, total_cash_cost,
        cash_out (= gross + employer_liability).
      - "reported_totals": the totals row as printed by ADP (for reconciliation).

    Raises ValueError if the header cannot be located or if the summed rows do
    not reconcile to the totals ADP printed in the file.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    col = _find_header(ws)
    if col is None:
        raise ValueError(f"Could not locate payroll header row in {path!r}")

    name_idx = col["employee"]
    rows: List[Dict[str, Any]] = []
    reported: Optional[Dict[str, float]] = None

    for raw in ws.iter_rows(values_only=True):
        # The first cell of summary/section rows carries a label spanning the
        # row (e.g. "Company Totals and Count :  50"); skip the section header.
        first = raw[0]
        if isinstance(first, str) and first.startswith("Company Totals"):
            continue

        name = raw[name_idx] if name_idx < len(raw) else None

        # Skip the header row itself (its "Name" cell literally reads "Name").
        if isinstance(name, str) and name.strip() == "Name":
            continue

        # A data row has a person's name; the company-total row has a blank
        # name but populated numeric columns.
        is_data = isinstance(name, str) and name.strip() != ""
        gross_idx = col["gross"]
        has_gross = gross_idx < len(raw) and isinstance(raw[gross_idx], (int, float))

        if is_data:
            rec: Dict[str, Any] = {}
            for field, idx in col.items():
                val = raw[idx] if idx < len(raw) else None
                if field in _NUMERIC_FIELDS:
                    rec[field] = _to_float(val)
                elif field == "date":
                    rec[field] = str(val).strip() if val is not None else None
                else:
                    rec[field] = str(val).strip() if isinstance(val, str) else val
            rec["cash_out"] = rec["gross"] + rec["employer_liability"]
            rows.append(rec)
        elif reported is None and not is_data and has_gross:
            # Company totals row: blank name, populated numbers.
            reported = {
                field: _to_float(raw[idx]) if idx < len(raw) else 0.0
                for field, idx in col.items()
                if field in _NUMERIC_FIELDS
            }

    if not rows:
        raise ValueError(f"No payroll data rows found in {path!r}")

    summed = {
        "gross": round(sum(r["gross"] for r in rows), 2),
        "employee_tax": round(sum(r["employee_tax"] for r in rows), 2),
        "employer_liability": round(sum(r["employer_liability"] for r in rows), 2),
    }

    # Reconcile our row sums against ADP's printed company totals.
    if reported is not None:
        for key in ("gross", "employee_tax", "employer_liability"):
            if round(reported.get(key, 0.0), 2) != summed[key]:
                raise ValueError(
                    f"Parsed {key} ({summed[key]}) does not reconcile to ADP "
                    f"reported total ({round(reported.get(key, 0.0), 2)})"
                )

    gross = summed["gross"]
    employer_liability = summed["employer_liability"]
    employee_tax = summed["employee_tax"]
    cash_out = round(gross + employer_liability, 2)

    # total_cash_cost comes from the file (ADP's "Total Expenses" total) when
    # available; otherwise derive it as gross + employer liability.
    if reported is not None and "total_expense" in reported:
        total_cash_cost = round(reported["total_expense"], 2)
    else:
        total_cash_cost = cash_out

    totals = {
        "gross": gross,
        "employee_tax": employee_tax,
        "employer_liability": employer_liability,
        "total_cash_cost": total_cash_cost,
        "cash_out": cash_out,
    }

    return {
        "rows": rows,
        "totals": totals,
        "reported_totals": reported,
    }
